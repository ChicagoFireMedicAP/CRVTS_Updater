"""
ts_export.py — TeleStaff Export Automation for CRVTS

Replaces the 40-step manual Excel process documented in
"TeleStaff Export Procedures for the Vacancy Tracking System."

Takes two TeleStaff downloads (Assignment Report + People CSV),
joins and transforms them, and outputs TS EXP.xlsx ready to drop
into SharePoint for the CRVTS Power Query.

Requires: pip install pandas openpyxl requests
"""

import re
import sys
from datetime import datetime
from pathlib import Path
from tkinter import Tk, filedialog

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependencies. Run:\n  pip install pandas openpyxl")
    sys.exit(1)


OUTPUT_DIR = Path.home() / "Downloads"


# ── File Picker ────────────────────────────────────────────────────────────────

def pick_file(title, filetypes):
    """Pop a native file-open dialog and return the selected path (or empty string)."""
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=title, initialdir=str(OUTPUT_DIR), filetypes=filetypes
    )
    root.destroy()
    return path


# ── Assignment Report Parser ──────────────────────────────────────────────────
#
# The Assignment Report is now a real .xlsx file (as of the TeleStaff migration).
# Previously it was XML/SpreadsheetML disguised as .xls — that format is gone.
# Row 0 is a junk row with non-person data; we filter it out by checking Person.
#

def parse_assignment_report(filepath):
    """Parse the TeleStaff Assignment Report xlsx into a DataFrame.

    Returns a DataFrame with columns:
    Institution, Region, Station, Unit, Person, Employee ID, File,
    Shift, Daley, From, Rank
    """
    print(f"  Parsing: {Path(filepath).name}")

    df = pd.read_excel(filepath, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    # Drop junk rows — only keep rows with an actual person name
    df = df[df["Person"].notna() & df["Person"].astype(str).str.strip().astype(bool)].reset_index(drop=True)

    # "Rank (Qual)" column isn't used anywhere downstream
    if "Rank (Qual)" in df.columns:
        df.drop(columns=["Rank (Qual)"], inplace=True)

    print(f"  Headers: {df.columns.tolist()}")
    print(f"  Parsed {len(df)} records")

    # Quick sanity check so we catch column misalignment early
    if len(df) > 0:
        sample_person = str(df.iloc[0]["Person"])
        print(f"  Sanity check — first Person: {sample_person}")
        if "(" not in sample_person and sample_person:
            print("  ⚠ WARNING: Person column doesn't contain parenthetical — columns may be misaligned!")

    return df


# ── Column Derivation ─────────────────────────────────────────────────────────
#
# These columns are derived from existing data without any external calls:
#   Name         = Person with the parenthetical stripped out
#   TS Assignment= The text inside the parenthetical (unit/company code)
#   PLT          = Daley value if present, else EMS platoon from shift name, else 5
#

def extract_name(person):
    """'GROSZEK(515), STEPHEN F.' → 'GROSZEK, STEPHEN F.'"""
    if not person:
        return ""
    return re.sub(r"\([^)]*\)", "", person).strip()


def extract_assignment(person):
    """'GROSZEK(515), STEPHEN F.' → '515'"""
    if not person:
        return ""
    m = re.search(r"\(([^)]*)\)", person)
    return m.group(1) if m else ""


def compute_plt(daley, shift):
    """Replicate the PLT formula from the manual process (step 21).

    Priority: use Daley if it exists, otherwise check the shift name
    for 'EMS Platoon N' and return 'EMSN', otherwise default to 5.
    """
    if daley:
        return daley
    if shift:
        m = re.search(r"EMS\s+Platoon\s+(\d)", shift)
        if m:
            return f"EMS{m.group(1)}"
    return 5


def add_derived_columns(df):
    """Add Name, PLT, and TS Assignment to the base DataFrame."""
    df["Name"] = df["Person"].apply(extract_name)
    df["PLT"] = df.apply(lambda r: compute_plt(r.get("Daley", ""), r.get("Shift", "")), axis=1)
    df["TS Assignment"] = df["Person"].apply(extract_assignment)
    return df


# ── IDPH License Status ───────────────────────────────────────────────────────
#
# IDPH status is determined from the People CSV. We check Specialty first
# (structured comma-separated list), then fall back to the Name field
# which sometimes has it embedded like (*.EMTP/EMS3).
#
# The manual process did this with three separate filtered downloads and
# an INDEX/MATCH — this does the same thing in one pass.
#

def idph_from_specialty(specialty_str):
    """Parse IDPH license type from the Specialty field."""
    if not specialty_str or pd.isna(specialty_str):
        return "NONE"
    specs = [s.strip() for s in str(specialty_str).split(",")]
    # Order matters: EMTBP check first because .EMT is a substring of .EMTBP
    if ".EMTBP" in specs:
        return "EMT (PM Drop)"
    if ".EMTP" in specs:
        return "PAR"
    if any(s == ".EMT" for s in specs):
        return "EMT"
    return "NONE"


def idph_from_name(name_str):
    """Fallback: extract IDPH from name strings like 'SMITH(A6), JOE (*.EMTP/EMS3)'."""
    if not name_str or pd.isna(name_str):
        return "NONE"
    name = str(name_str)
    if "*.EMTBP" in name:
        return "EMT (PM Drop)"
    if "*.EMTP" in name:
        return "PAR"
    if "*.EMT" in name:
        return "EMT"
    return "NONE"


# ── People CSV Enrichment ─────────────────────────────────────────────────────
#
# The People CSV (exported via TeleStaff > People > gear > Export People CSV)
# provides two things we can't get from the Assignment Report:
#   - Promoted date (from the "Promoted" column)
#   - IDPH license status (derived from "Specialty" and/or "Name" columns)
#
# We join on Payroll ID (People CSV) = File (Assignment Report).
#

def find_column(df, candidates):
    """Find a column name case-insensitively from a list of possible names."""
    for candidate in candidates:
        for col in df.columns:
            if col.lower().strip() == candidate.lower().strip():
                return col
    return None


def load_people_file(filepath):
    """Load the TeleStaff People export (CSV or XLSX)."""
    print(f"  Loading: {Path(filepath).name}")
    ext = Path(filepath).suffix.lower()
    if ext == ".csv":
        df = pd.read_csv(filepath, encoding="utf-8-sig")
    elif ext in (".xlsx", ".xls"):
        df = pd.read_excel(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}")
    df.columns = [c.strip() for c in df.columns]
    print(f"  {len(df)} records loaded")
    return df


def enrich_from_people(df_main, df_people):
    """Add Promoted and IDPH Status columns using the People export.

    Joins on File (Assignment Report) = Payroll ID (People CSV).
    """
    payroll_col = find_column(df_people, ["Payroll ID", "PayrollID", "Payroll_ID", "payrollId"])
    if not payroll_col:
        print(f"  ⚠ No Payroll ID column found in People file.")
        print(f"    Available columns: {df_people.columns.tolist()}")
        df_main["Promoted"] = ""
        df_main["IDPH Status"] = "NONE"
        return df_main

    promo_col = find_column(df_people, ["Promoted"])
    spec_col = find_column(df_people, ["Specialty"])
    name_col = find_column(df_people, ["Name"])
    print(f"  Mapping columns: payroll={payroll_col}, promoted={promo_col}, "
          f"specialty={spec_col}, name={name_col}")

    # Build lookup dicts keyed by payroll ID string
    df_people["_pid"] = df_people[payroll_col].astype(str).str.strip()
    promo_lookup = {}
    idph_lookup = {}

    for _, row in df_people.iterrows():
        pid = row["_pid"]
        if not pid or pid == "nan":
            continue

        # Promoted date — keep whatever format the CSV gives us
        if promo_col and pd.notna(row.get(promo_col)):
            val = row[promo_col]
            if isinstance(val, (datetime, pd.Timestamp)):
                promo_lookup[pid] = val.strftime("%Y-%m-%d")
            elif val:
                promo_lookup[pid] = str(val)

        # IDPH: structured Specialty field first, name string as fallback
        status = "NONE"
        if spec_col:
            status = idph_from_specialty(row.get(spec_col, ""))
        if status == "NONE" and name_col:
            status = idph_from_name(row.get(name_col, ""))
        idph_lookup[pid] = status

    # Apply lookups — File column in Assignment Report = Payroll ID
    df_main["_fstr"] = df_main["File"].astype(str).str.strip()
    df_main["Promoted"] = df_main["_fstr"].map(promo_lookup).fillna("")
    df_main["IDPH Status"] = df_main["_fstr"].map(idph_lookup).fillna("NONE")
    df_main.drop(columns=["_fstr"], inplace=True)

    # Report match rates so you know if something's off
    total = len(df_main)
    idph_matched = df_main["IDPH Status"].ne("NONE").sum()
    promo_filled = df_main["Promoted"].ne("").sum()
    print(f"  IDPH matched: {idph_matched}/{total} ({idph_matched / total * 100:.1f}%)")
    print(f"  Promoted filled: {promo_filled}/{total} ({promo_filled / total * 100:.1f}%)")

    return df_main


# ── Excel Output ──────────────────────────────────────────────────────────────
#
# Writes a 4-sheet workbook matching the format CRVTS Power Query expects:
#   TS Assign   — base data + "Text Between Delimiters" (replaces Power Query extraction)
#   TS EXP      — the main sheet CRVTS reads from
#   TS Promoted — full People export for reference/INDEX-MATCH
#   IDPH Lic    — payroll ID + license type (no headers), for reference
#

def add_table(ws, table_name, num_rows, num_cols):
    """Format a worksheet range as an Excel table."""
    if num_rows < 1:
        return
    ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
    safe_name = re.sub(r"[^A-Za-z0-9_]", "", table_name)
    t = Table(displayName=safe_name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(t)


def auto_width(ws):
    """Auto-fit column widths based on content (capped at 40 chars)."""
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def write_workbook(df_main, df_people, output_path):
    """Write the final TS_EXP.xlsx with all four sheets."""
    print(f"\nWriting workbook...")
    wb = Workbook()

    # ── TS Assign ──
    ws1 = wb.active
    ws1.title = "TS Assign"
    cols1 = [
        "Institution", "Region", "Station", "Unit", "Person", "Name",
        "Employee ID", "File", "Shift", "Daley", "PLT", "From",
        "Promoted", "Rank", "Text Between Delimiters",
    ]
    ws1.append(cols1)
    for _, r in df_main.iterrows():
        ws1.append([
            r.get("Institution", ""), r.get("Region", ""), r.get("Station", ""),
            r.get("Unit", ""), r.get("Person", ""), r.get("Name", ""),
            r.get("Employee ID", ""), r.get("File", ""), r.get("Shift", ""),
            r.get("Daley", ""), r.get("PLT", ""), r.get("From", ""),
            r.get("Promoted", ""), r.get("Rank", ""), r.get("TS Assignment", ""),
        ])
    add_table(ws1, "TSAssign", len(df_main), len(cols1))
    auto_width(ws1)

    # ── TS EXP (this is what CRVTS Power Query reads) ──
    ws2 = wb.create_sheet("TS EXP")
    cols2 = [
        "Institution", "Region", "Station", "Unit", "Person", "Name",
        "Employee ID", "File", "Shift", "Daley", "PLT", "From",
        "Promoted", "IDPH Status", "Rank", "TS Assignment",
    ]
    ws2.append(cols2)
    for _, r in df_main.iterrows():
        ws2.append([
            r.get("Institution", ""), r.get("Region", ""), r.get("Station", ""),
            r.get("Unit", ""), r.get("Person", ""), r.get("Name", ""),
            r.get("Employee ID", ""), r.get("File", ""), r.get("Shift", ""),
            r.get("Daley", ""), r.get("PLT", ""), r.get("From", ""),
            r.get("Promoted", ""), r.get("IDPH Status", ""), r.get("Rank", ""),
            r.get("TS Assignment", ""),
        ])
    add_table(ws2, "TSEXP", len(df_main), len(cols2))
    auto_width(ws2)

    # ── TS Promoted (full People export, preserved for reference) ──
    ws3 = wb.create_sheet("TS Promoted")
    if df_people is not None and len(df_people) > 0:
        pcols = list(df_people.columns)
        ws3.append(pcols)
        for _, r in df_people.iterrows():
            ws3.append([
                r.get(c, "") if pd.notna(r.get(c, "")) else "" for c in pcols
            ])
        add_table(ws3, "TSPromoted", len(df_people), len(pcols))
    else:
        ws3.append(["No People data loaded"])
    auto_width(ws3)

    # ── IDPH Lic (no headers — just payroll ID + license type) ──
    ws4 = wb.create_sheet("IDPH Lic")
    for _, r in df_main.iterrows():
        status = r.get("IDPH Status", "NONE")
        if status != "NONE":
            fv = r.get("File", "")
            try:
                fv = int(fv)
            except (ValueError, TypeError):
                pass
            ws4.append([fv, status])

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ── Validation ─────────────────────────────────────────────────────────────────

def validate(df):
    """Print a summary so you can eyeball things before uploading."""
    print("\n── Validation ──")
    print(f"  Total rows: {len(df)}")

    print(f"\n  IDPH Status:")
    for val, cnt in df["IDPH Status"].value_counts().items():
        print(f"    {val}: {cnt}")

    print(f"\n  PLT (top 10):")
    for val, cnt in df["PLT"].value_counts().head(10).items():
        print(f"    {val}: {cnt}")

    ranked = df[df["Rank"].astype(str).str.strip().ne("")]
    print(f"\n  Rank filled: {len(ranked)}/{len(df)}")
    if len(ranked):
        for val, cnt in ranked["Rank"].value_counts().head(10).items():
            print(f"    {val}: {cnt}")

    promo_filled = df["Promoted"].astype(str).str.strip().ne("").sum()
    print(f"\n  Promoted filled: {promo_filled}/{len(df)}")

    print(f"\n  Sample rows:")
    for i, (_, r) in enumerate(df.head(3).iterrows()):
        print(
            f"    [{i + 1}] Person={r['Person']} | Inst={r['Institution']} | "
            f"Region={r.get('Region', '')} | Shift={r['Shift']} | "
            f"Daley={r.get('Daley', '')} | PLT={r['PLT']} | "
            f"IDPH={r['IDPH Status']} | Rank={r.get('Rank', '')} | "
            f"TS Assign={r.get('TS Assignment', '')}"
        )


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  TeleStaff Export Automation for CRVTS")
    print("=" * 60)

    # Step 1: Assignment Report
    print("\n[Step 1] Select the Assignment Report file...")
    report_path = pick_file(
        "Select TeleStaff Assignment Report",
        [("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    if not report_path:
        print("  Cancelled.")
        return

    # Step 2: Parse the report and derive what we can without the People file
    print("\n[Step 2] Parsing Assignment Report...")
    df = parse_assignment_report(report_path)
    df = add_derived_columns(df)

    # Step 3: People CSV for Promoted + IDPH
    print("\n[Step 3] Select the People CSV export...")
    people_path = pick_file(
        "Select TeleStaff People Export (CSV or XLSX)",
        [("CSV/Excel files", "*.csv *.xlsx *.xls"), ("All files", "*.*")],
    )
    df_people = None
    if people_path:
        df_people = load_people_file(people_path)
        df = enrich_from_people(df, df_people)
    else:
        print("  No People file selected — Promoted and IDPH will be empty.")
        df["Promoted"] = ""
        df["IDPH Status"] = "NONE"

    # Step 4: Choose save location and write the workbook
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    output_path = filedialog.asksaveasfilename(
        title="Save TS EXP workbook",
        initialdir=str(OUTPUT_DIR),
        initialfile="TS EXP.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    if not output_path:
        print("  Save cancelled.")
        return
    output_path = Path(output_path)

    write_workbook(df, df_people, output_path)

    # Step 5: Show summary
    validate(df)

    print(f"\n{'=' * 60}")
    print(f"  DONE!")
    print(f"  Output: {output_path}")
    print(f"  Next: Drag to SharePoint → Refresh CRVTS Power Query")
    print(f"{'=' * 60}")
    input("\nPress Enter to close...")


if __name__ == "__main__":
    main()
